import logging
from functools import cached_property
from os import PathLike
from typing import List, Optional

from buildingmotif.ingresses.base import Record, RecordIngressHandler

try:
    from openpyxl import load_workbook
except ImportError:
    logging.critical(
        "Install the 'xlsx-ingress' module, e.g. 'pip install buildingmotif[xlsx-ingress]'"
    )


class XLSXIngress(RecordIngressHandler):
    """Reads sheets from a XLSX file and exposes them as records. The 'rtype'
    field of each Record gives the name of the sheet.
    """

    def __init__(self, filename: PathLike, limit: Optional[int] = -1):
        """
        Path to the .xlsx file to be ingested

        :param filename: Path to a .xlsx file
        :type filename: PathLike
        :param limit: The maximum number of rows to read from each sheet. If -1 (default), reads all rows.
        :type limit: Optional[int], optional
        """

        self.filename = filename
        self.limit = limit or -1

    @cached_property
    def records(self) -> List[Record]:
        """The set of rows in all sheets in the XLSX file.

        :return: A Record representing a row in a sheet. The sheetname is stored
                in the 'rtype' field. The 'fields' field contains key-value pairs
                for each row; the keys are the names of the columns and the values
                are the cell values at that column for the given row.
        :rtype: List[Record]
        """
        records = []
        # using data_only means that the cells will contain data, rather than a formula.
        wb = load_workbook(self.filename, data_only=True)  # noqa
        for sheetname in wb.sheetnames:
            sheet = wb[sheetname]
            columns = [sheet.cell(1, c + 1).value for c in range(sheet.max_column)]
            upper_range = sheet.max_row + 1 if self.limit < 0 else self.limit
            for row in range(2, upper_range):
                fields = {
                    columns[c]: sheet.cell(row, c + 1).value
                    for c in range(sheet.max_column)
                }
                records.append(
                    Record(
                        rtype=sheetname,
                        fields=fields,
                    )
                )
        return records
