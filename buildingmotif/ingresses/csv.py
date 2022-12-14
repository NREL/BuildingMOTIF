from csv import DictReader
from functools import cached_property
from pathlib import Path
from typing import List

from openpyxl import load_workbook

from buildingmotif.ingresses.base import Record, RecordIngressHandler


class CSVIngress(RecordIngressHandler):
    """Reads rows from a CSV file and exposes them as records.
    The type of the record is the name of the CSV file
    """

    def __init__(self, filename: Path):
        self.filename = filename

    @cached_property
    def records(self) -> List[Record]:
        records = []
        rdr = DictReader(open(self.filename))
        for row in rdr:
            rec = Record(
                rtype=str(self.filename),
                fields=row,
            )
            records.append(rec)

        return records


class XLSXIngress(RecordIngressHandler):
    """Reads sheets from a XLSX file and exposes them as records. The 'rtype'
    field of each Record gives the name of the sheet.
    """

    def __init__(self, filename: Path):
        """
        Path to the .xlsx file to be ingested

        :param filename: Path to a .xlsx file
        :type filename: Path
        """
        self.filename = filename

    @cached_property
    def records(self) -> List[Record]:
        """The set of rows in all sheets in the XLSX file.

        :return: A Record representing a row in a sheet. The sheetname is stored
                in the 'rtype' field. The 'fields' field contains key-value pairs
                for each row; the keys are the names of the columns and the values
                are the cell values at that column for the given row.
        :rtype: List[Record]
        """
        records = []
        wb = load_workbook(self.filename)
        for sheetname in wb.sheetnames:
            sheet = wb[sheetname]
            columns = [sheet.cell(1, c + 1).value for c in range(sheet.max_column)]
            for row in range(2, sheet.max_row + 1):
                fields = {
                    columns[c]: sheet.cell(row, c + 1).value
                    for c in range(sheet.max_column)
                }
                records.append(
                    Record(
                        rtype=sheetname,
                        fields=fields,
                    )
                )
        return records
