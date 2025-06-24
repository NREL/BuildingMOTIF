import csv
import logging
import warnings
from collections import Counter
from copy import copy
from dataclasses import dataclass
from io import BytesIO, StringIO
from itertools import chain
from os import PathLike
from secrets import token_hex
from typing import (
    TYPE_CHECKING,
    Dict,
    Generator,
    List,
    Optional,
    Set,
    Tuple,
    Union,
    overload,
)

import rdflib
from rdflib.term import Node

from buildingmotif import get_building_motif
from buildingmotif.database.errors import TemplateDependencyNotFound, TemplateNotFound
from buildingmotif.dataclasses.model import Model
from buildingmotif.namespaces import bind_prefixes
from buildingmotif.template_matcher import Mapping, TemplateMatcher
from buildingmotif.utils import (
    PARAM,
    _strip_param,
    combine_graphs,
    copy_graph,
    remove_triples_with_node,
    replace_nodes,
)

if TYPE_CHECKING:
    from buildingmotif import BuildingMOTIF
    from buildingmotif.dataclasses.library import Library


@dataclass
class Template:
    """This class mirrors :py:class`database.tables.DBTemplate`."""

    _id: int
    _name: str
    body: rdflib.Graph
    optional_args: List[str]
    variadic_args: List[str]
    _bm: "BuildingMOTIF"

    @classmethod
    def load(cls, id: int) -> "Template":
        """Load template from database.

        :param id: id of template
        :type id: int
        :return: loaded template
        :rtype: Template
        """
        bm = get_building_motif()
        db_template = bm.table_connection.get_db_template(id)
        body = bm.graph_connection.get_graph(db_template.body_id)

        return cls(
            _id=db_template.id,
            _name=db_template.name,
            optional_args=db_template.optional_args,
            variadic_args=db_template.variadic_args,
            body=body,
            _bm=bm,
        )

    def in_memory_copy(self) -> "Template":
        """Copy this template.

        :return: copy of this template
        :rtype: Template
        """
        return Template(
            _id=-1,
            _name=self._name,
            body=copy_graph(self.body, preserve_blank_nodes=False),
            optional_args=self.optional_args[:],
            variadic_args=self.variadic_args[:],
            _bm=self._bm,
        )

    @property
    def id(self):
        return self._id

    @id.setter
    def id(self, new_id):
        raise AttributeError("Cannot modify db id")

    @property
    def name(self):
        return self._name

    @name.setter
    def name(self, new_name: str) -> None:
        self._bm.table_connection.update_db_template_name(self._id, new_name)
        self._name = new_name

    def get_dependencies(self) -> Tuple["Dependency", ...]:
        """Get the template's dependencies.

        :return: a tuple of dependencies
        :rtype: Tuple
        """
        return tuple(
            [
                Dependency.load(dep.id)
                for dep in self._bm.table_connection.get_db_template_dependencies(
                    self._id
                )
            ]
        )

    @overload
    def add_dependency(self, dependency: "Template", args: Dict[str, str]) -> None:
        """Add dependency to template.

        :param dependency: dependency to add
        :type dependency: Template
        :param args: dictionary of dependency arguments
        :type args: Dict[str, str]
        """

    @overload
    def add_dependency(
        self, dependency_library: str, dependency_template: str, args: Dict[str, str]
    ) -> None:
        """Add dependency to template.

        :param dependency_library: name of library containing depdency
        :type dependency_library: str
        :param dependency_template: name of template depedency
        :type dependency_template: str
        :param args: dictionary of dependency arguments
        :type args: Dict[str, str]
        """

    def add_dependency(self, *args, **kwargs):
        total_args = len(args) + len(kwargs)
        if total_args == 2:
            dependency: "Template" = kwargs.get("dependency", args[0])
            args: Dict[str, str] = kwargs.get("args", args[1])
            self._bm.table_connection.add_template_dependency_preliminary(
                self.id, dependency.defining_library.name, dependency.name, args
            )
        elif total_args == 3:
            dependency_library: str = kwargs.get("dependency_library", args[0])
            dependency_template: str = kwargs.get("dependency_template", args[1])
            args: Dict[str, str] = kwargs.get("args", args[2])
            self._bm.table_connection.add_template_dependency_preliminary(
                self.id, dependency_library, dependency_template, args
            )

    def check_dependencies(self):
        """
        Verifies that all dependencies have valid references to the parameters
        in the dependency or (recursively) its dependencies. Raises an exception if any
        issues are found.

        It is recommended to call this after *all* templates in a library have been
        loaded in to the DB, else this might falsely raise an error.
        """
        for dep in self._bm.table_connection.get_db_template_dependencies(self.id):
            self._bm.table_connection.check_template_dependency_relationship(dep)

    def remove_dependency(self, dependency: "Template") -> None:
        """Remove dependency from template.

        :param dependency: dependency to remove
        :type dependency: Template
        """
        self._bm.table_connection.delete_template_dependency(self.id, dependency.id)

    @property
    def all_parameters(self, error_on_missing_dependency: bool = True) -> Set[str]:
        """The set of all parameters used in this template *including* its
        dependencies. Includes optional parameters.

        :param error_on_missing_dependency: Raise an erorr if a template has a missing depedency
        :type error_on_missing_dependency: bool
        :return: set of parameters *with* dependencies
        :rtype: Set[str]
        """
        # handle local parameters first
        params = set(self.parameters)

        # then handle dependencies
        for dep in self.get_dependencies():
            if dep.template is None:
                if error_on_missing_dependency:
                    raise TemplateNotFound(name=dep.dependency_template_name)
                continue
            params.update(dep.template.parameters)
        return params

    @property
    def parameters(self) -> Set[str]:
        """The set of all parameters used in this template *excluding* its
        dependencies. Includes optional parameters.

        :return: set of parameters *without* dependencies
        :rtype: Set[str]
        """
        # handle local parameters first
        nodes = chain.from_iterable(self.body.triples((None, None, None)))
        params = {str(p)[len(PARAM) :] for p in nodes if str(p).startswith(PARAM)}
        return params

    @property
    def dependency_parameters(
        self, error_on_missing_dependency: bool = True
    ) -> Set[str]:
        """The set of all parameters used in this template's dependencies, including
        optional parameters.

        :param error_on_missing_dependency: Raise an erorr if a template has a missing depedency
        :type error_on_missing_dependency: bool
        :return: set of parameters used in dependencies
        :rtype: Set[str]
        """
        params: Set[str] = set()
        for dep in self.get_dependencies():
            if dep.template is None:
                if error_on_missing_dependency:
                    raise TemplateNotFound(name=dep.dependency_template_name)
                continue
            params = params.union(dep.template.parameters)
        return params

    @property
    def parameter_counts(self, error_on_missing_dependency: bool = True) -> Counter:
        """An addressable histogram of the parameter name counts in this
        template and all of its transitive dependencies.

        :param error_on_missing_dependency: Raise an erorr if a template has a missing depedency
        :type error_on_missing_dependency: bool
        :return: count of parameters
        :rtype: Counter
        """
        counts: Counter = Counter()
        counts.update(self.parameters)
        for dep in self.get_dependencies():
            if dep.template is None:
                if error_on_missing_dependency:
                    raise TemplateNotFound(name=dep.dependency_template_name)
                continue
            counts.update(dep.template.parameter_counts)
        return counts

    # TODO: method to get the 'types' of the parameters

    def dependency_for_parameter(self, param: str) -> Optional["Template"]:
        """Returns the dependency that uses the given parameter if one exists.

        :param param: parameter to search for
        :type param: str
        :return: dependency that uses the given parameter
        :rtype: Optional[Template]
        """
        for dep in self.get_dependencies():
            if param in dep.args.values():
                return dep.template
        return None

    def to_inline(self, preserve_args: Optional[List[str]] = None) -> "Template":
        """Return an inline-able copy of this template.

        Suffixes all parameters with a unique identifier that will avoid
        parameter name collisions when templates are combined with one another.
        Any argument names in the `preserve_args` list will not be adjusted.

        :param preserve_args: parameters whose names will be preserved,
            defaults to None
        :type preserve_args: Optional[List[str]], optional
        :return: a template with globally unique parameters
        :rtype: Template
        """
        templ = self.in_memory_copy()
        suffix = f"{token_hex(4)}-inlined"
        # the lookup table of old to new parameter names
        to_replace = {}
        for param in templ.parameters:
            # skip if (a) we want to preserve the param or (b) it is already inlined
            if (preserve_args and param in preserve_args) or (
                param.endswith("-inlined")
            ):
                continue
            param = PARAM[param]
            to_replace[param] = rdflib.URIRef(f"{param}-{suffix}")
        replace_nodes(templ.body, to_replace)
        return templ

    @property
    def transitive_parameters(self) -> Set[str]:
        """Get all parameters used in this template and its dependencies.

        :return: set of all parameters
        :rtype: Set[str]
        """
        params = set(self.parameters)
        for dep in self.get_dependencies():
            if dep.template is None:
                raise TemplateNotFound(name=dep.dependency_template_name)
            transitive_params = dep.template.transitive_parameters
            rename_params: Dict[str, str] = {
                ours: theirs for ours, theirs in dep.args.items()
            }
            name_prefix = dep.args.get("name")
            for param in transitive_params:
                if param not in dep.args and param != "name":
                    rename_params[param] = f"{name_prefix}-{param}"
            params.update(rename_params.values())
        return params

    def inline_dependencies(self) -> "Template":
        """Copies this template with all dependencies recursively inlined.

        Parameters of dependencies will be renamed to avoid confusion.

        :return: copy of this template with all dependencies inlined
        :rtype: Template
        """
        templ = self.in_memory_copy()
        # if this template has no dependencies, then return unaltered
        if not self.get_dependencies():
            return templ

        # start with this template's parameters; this recurses into each
        # dependency to inline dependencies
        for dep in self.get_dependencies():
            if dep.template is None:
                raise TemplateNotFound(name=dep.dependency_template_name)
            # get the inlined version of the dependency
            deptempl = dep.template.inline_dependencies()

            # replace dependency parameters with the names they inherit
            # through the provided bindings
            rename_params: Dict[str, str] = {
                ours: _strip_param(theirs) for ours, theirs in dep.args.items()
            }
            # replace all parameters *not* mentioned in the args by prefixing
            # them with the 'name' parameter binding; this is guaranteed
            # to exist
            name_prefix = _strip_param(dep.args["name"])
            # for each parameter in the dependency...
            for param in deptempl.parameters:
                # if it does *not* have a mapping in the dependency, then
                # prefix the parameter with the value of the 'name' binding
                # to scope it properly
                param = _strip_param(param)
                if param not in dep.args and param != "name":
                    rename_params[param] = f"{name_prefix}-{param}"

            # replace the parameters in the dependency template
            replace_nodes(
                deptempl.body, {PARAM[k]: PARAM[v] for k, v in rename_params.items()}
            )
            # rename the optional_args in the dependency template too
            deptempl.optional_args = [
                rename_params.get(arg, arg) for arg in deptempl.optional_args
            ]

            # at this point, deptempl's parameters are all unique with respect to
            # the parent template. They are either renamed explicitly via the dependency's
            # args or implicitly via prefixing with the 'name' parameter.

            # Next, we need to determine which of deptempl's parameters are optional
            # and add these to the parent template's optional_args list.

            # get the parent template's optional args
            templ_optional_args = set(templ.optional_args)

            # represents the optional parameters of the dependency template
            deptempl_opt_args: Set[str] = set()

            # these optional parameters come from two places.
            # 1. the dependency template itself (its optional_args)
            deptempl_opt_args.update(deptempl.optional_args)
            # 1a. remove any parameters that have the same name as a parameter in the
            #     parent but are not optional in the parent
            deptempl_opt_args.difference_update(templ.parameters)
            # 2. having the same name as an optional parameter in the parent template
            #    (templ_optional_args)
            deptempl_opt_args.update(
                templ_optional_args.intersection(deptempl.parameters)
            )
            # 2a. if the 'name' of the deptempl is optional (given by the parent template),
            #   then all the arguments inside deptempl become optional
            #   (deptempl.parameters)
            if rename_params["name"] in deptempl_opt_args:
                # mark all of deptempl's parameters as optional
                deptempl_opt_args.update(deptempl.parameters)

            # convert our set of optional params to a list and assign to the parent template
            # 1. get required parameters from the original template
            # 2. calculate all optional requirements from the dependency template and the original template
            # 3. remove required parameters from the optional requirements
            # This avoids a bug where an optional dependency makes reference to a required parameter, and then
            # subsequent inlining of the dependency without optional args would remove the required parameter
            required = templ.parameters - templ_optional_args
            templ.optional_args = list(
                templ_optional_args.union(deptempl_opt_args) - required
            )

            # append the inlined template into the parent's body
            templ.body += deptempl.body

        return templ

    def evaluate(
        self,
        bindings: Dict[str, Node],
        namespaces: Optional[Dict[str, rdflib.Namespace]] = None,
        require_optional_args: bool = False,
        warn_unused: bool = True,
    ) -> Union["Template", rdflib.Graph]:
        """Evaluate the template with the provided bindings.

        If all parameters in the template have a provided binding, then a graph
        will be returned. Otherwise, a new Template will be returned that
        incorporates the provided bindings and preserves unbound parameters. If
        `require_optional_args` is True, then the template evaluation will not
        return a graph unless all optional arguments are bound. If
        `require_optional_args` is False, then the template evaluation will
        return a graph even if some optional arguments are unbound.

        :param bindings: map of parameter {name: RDF term} to substitute
        :type bindings: Dict[str, Node]
        :param namespaces: namespace bindings to add to the graph,
            defaults to None
        :type namespaces: Optional[Dict[str, rdflib.Namespace]], optional
        :param require_optional_args: whether to require all optional arguments
            to be bound, defaults to False
        :type require_optional_args: bool
        :param warn_unused: if True, print a warning if there were any parameters left
            unbound during evaluation. If 'require_optional_args' is True,
            then this will consider optional parameters when producing the warning;
            otherwise, optional paramters will be ignored; defaults to True
        :type warn_unused: bool
        :return: either a template or a graph, depending on whether all
            parameters were provided
        :rtype: Union[Template, rdflib.Graph]
        """
        # TODO: handle datatype properties
        templ = self.in_memory_copy()
        # put all of the parameter names into the PARAM namespace so they can be
        # directly subsituted in the template body
        uri_bindings: Dict[Node, Node] = {PARAM[k]: v for k, v in bindings.items()}
        # replace the param:<name> URIs in the template body with the bindings
        # provided in the call to evaluate()
        replace_nodes(templ.body, uri_bindings)
        leftover_params = (
            templ.parameters.difference(bindings.keys())
            if not require_optional_args
            else (templ.parameters.union(self.optional_args)).difference(
                bindings.keys()
            )
        )
        # true if all parameters are now bound or only optional args are unbound
        if len(templ.parameters) == 0 or (
            not require_optional_args
            and templ.parameters.issubset(set(self.optional_args))
        ):
            bind_prefixes(templ.body)
            if namespaces:
                for prefix, namespace in namespaces.items():
                    templ.body.bind(prefix, namespace)
            if not require_optional_args:
                # remove all triples that touch unbound optional_args
                unbound_optional_args = set(templ.optional_args) - set(bindings.keys())
                for arg in unbound_optional_args:
                    remove_triples_with_node(templ.body, PARAM[arg])
            return templ.body
        if len(leftover_params) > 0 and warn_unused:
            warnings.warn(
                f"Parameters \"{', '.join(leftover_params)}\" were not provided during evaluation",
                UserWarning,
            )
        return templ

    def fill(
        self, ns: rdflib.Namespace, include_optional: bool = False
    ) -> Tuple[Dict[str, Node], rdflib.Graph]:
        """Evaluates the template with autogenerated bindings within the given
        namespace.

        :param ns: namespace to contain the autogenerated entities
        :type ns: rdflib.Namespace
        :param include_optional: if True, invent URIs for optional parameters; ignore if False
        :type include_optional: bool
        :return: a tuple of the bindings used and the resulting graph
        :rtype: Tuple[Dict[str, Node], rdflib.Graph]
        """
        bindings: Dict[str, Node] = {
            param: ns[f"{param}_{token_hex(4)}"]
            for param in self.parameters
            if include_optional or param not in self.optional_args
        }
        res = self.evaluate(bindings, require_optional_args=include_optional)
        assert isinstance(res, rdflib.Graph)
        return bindings, res

    @property
    def defining_library(self) -> "Library":
        """The library defining this template.

        :return: library
        :rtype: Library
        """
        from buildingmotif.dataclasses.library import Library

        return Library.load(
            self._bm.table_connection.get_library_defining_db_template(self.id).id
        )

    def library_dependencies(
        self, error_on_missing_dependency: bool = True
    ) -> List["Library"]:
        """Get library dependencies for this template.

        :param error_on_missing_dependency: Raise an erorr if a template has a missing depedency
        :type error_on_missing_dependency: bool
        :return: list of libraries
        :rtype: List[Library]
        """
        from buildingmotif.dataclasses.library import Library

        libs = {self.defining_library.id}
        for dep in self.get_dependencies():
            if dep.template is None:
                if error_on_missing_dependency:
                    raise TemplateNotFound(name=dep.dependency_template_name)
                continue
            libs.add(dep.template.defining_library.id)
        return [Library.load(id) for id in libs]

    def find_subgraphs(
        self, model: Model, *ontologies: rdflib.Graph
    ) -> Generator[Tuple[Mapping, rdflib.Graph, Optional["Template"]], None, None]:
        """Produces an iterable of subgraphs in the model that are partially or
        entirely covered by the provided template.

        :yield: iterable of subgraphs in the model
        :rtype: Generator[Tuple[Mapping, rdflib.Graph, Optional[Template]], None, None]
        """
        # TODO: can we figure out what ontology to use automatically?
        # if ontology is not specified, pull in all shapes related to this template's library
        # and all of its dependencies
        if len(ontologies) == 0:
            ontology = rdflib.Graph()
            for lib in self.library_dependencies(False):
                ontology += lib.get_shape_collection().graph
        else:
            ontology = combine_graphs(*ontologies)

        matcher = TemplateMatcher(model.graph, self, ontology)
        for mapping, sg in matcher.building_mapping_subgraphs_iter():
            yield mapping, sg, matcher.remaining_template(mapping)

    def generate_csv(self, path: Optional[PathLike] = None) -> Optional[StringIO]:
        """
        Generate a CSV for this template which contains a column for each template parameter.
        Once filled out, the resulting CSV file can be passed to a Template Ingress to populate a model.
        Returns a 'io.BytesIO' object which can be written to a file or sent to another program/function.

        :param path: if not None, writes the CSV to the indicated file
        :type path: PathLike, optional
        :return: String buffer containing the resulting CSV file
        :rtype: StringIO
        """
        all_parameters = copy(self.parameters)
        mandatory_parameters = all_parameters - set(self.optional_args)
        row_data = list(mandatory_parameters) + list(self.optional_args)

        if path is not None:
            # write directly to file
            with open(path, "w") as f:
                writer = csv.writer(f)
                writer.writerow(row_data)
            return None

        # write to in-memory file
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(row_data)
        return output

    def generate_spreadsheet(
        self, path: Optional[PathLike] = None
    ) -> Optional[BytesIO]:
        """
        Generate a spreadsheet for this template which contains a column for each template parameter.
        Once filled out, the resulting spreadsheet can be passed to a Template Ingress to populate a model.
        Returns a 'io.BytesIO' object which can be written to a file or sent to another program/function.

        :param path: if not None, writes the CSV to the indicated file
        :type path: PathLike, optional
        :return: Byte buffer containing the resulting spreadsheet file
        :rtype: BytesIO
        """
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            from openpyxl.worksheet.table import Table, TableStyleInfo
        except ImportError:
            logging.critical(
                "Install the 'xlsx-ingress' module, e.g. 'pip install buildingmotif[xlsx-ingress]'"
            )
            return None
        all_parameters = copy(self.parameters)
        mandatory_parameters = all_parameters - set(self.optional_args)

        workbook = Workbook()
        sheet = workbook.active
        if sheet is None:
            raise Exception("Could not open active sheet in Workbook")

        row_data = list(mandatory_parameters) + list(self.optional_args)
        for column_index, cell_value in enumerate(row_data, 1):
            column_letter = get_column_letter(column_index)
            sheet[f"{column_letter}1"] = cell_value  # type: ignore
            # Adjust column width based on cell content
            column_dimensions = sheet.column_dimensions[column_letter]  # type: ignore
            column_dimensions.width = max(column_dimensions.width, len(str(cell_value)))

        tab = Table(
            displayName="Table1", ref=f"A1:{get_column_letter(len(row_data))}10"
        )
        style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        tab.tableStyleInfo = style
        sheet.add_table(tab)

        if path is not None:
            # write directly to file
            workbook.save(path)
            return None

        # save the file in-memory and return the resulting buffer
        f = BytesIO()
        workbook.save(f)
        return f


@dataclass
class Dependency:
    """Dependency"""

    _id: int
    _bm: "BuildingMOTIF"

    _dependency_library_name: str
    _dependency_template_name: str
    _args: Dict[str, str]

    @classmethod
    def load(cls, id: int) -> "Dependency":
        bm = get_building_motif()
        dep = bm.table_connection.get_db_template_dependency(id)
        if dep is None:
            raise TemplateDependencyNotFound(id)

        return cls(
            _id=id,
            _bm=bm,
            _dependency_library_name=dep.dependency_library_name,
            _dependency_template_name=dep.dependency_template_name,
            _args=dep.args,
        )

    @property
    def dependency_template_name(self) -> str:
        return self._dependency_template_name

    @property
    def dependency_library_name(self) -> str:
        return self._dependency_library_name

    @property
    def args(self) -> Dict[str, str]:
        return self._args

    @property
    def template(self) -> Optional[Template]:
        db_dependency = self._bm.table_connection.get_db_template_dependency(self._id)
        if db_dependency is None:
            return None
        db_template = db_dependency.dependency_template
        if db_template is None:
            return None

        return Template.load(db_template.id)
