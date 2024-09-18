import csv
from io import StringIO
from pathlib import Path
from tempfile import NamedTemporaryFile

import openpyxl
from rdflib import Namespace
from rdflib.compare import isomorphic

from buildingmotif.dataclasses import Library
from buildingmotif.ingresses.csvingress import CSVIngress
from buildingmotif.ingresses.template import TemplateIngress
from buildingmotif.ingresses.xlsx import XLSXIngress

BLDG = Namespace("urn:bldg/")


# utility function for spreadsheet tests
def _add_spreadsheet_row(sheet, bindings):
    num_columns = sheet.max_column
    for column in range(1, num_columns + 1):
        param = sheet.cell(row=1, column=column).value
        sheet.cell(row=2, column=column).value = bindings[param][len(BLDG) :]


def _add_csv_row(params, tempfile, bindings):
    params = [param.strip() for param in params]
    if isinstance(tempfile, StringIO):
        w = csv.writer(tempfile)
        w.writerow([bindings[param][len(BLDG) :] for param in params])
        tempfile.flush()
    else:
        with open(Path(tempfile.name), "a") as f:
            w = csv.writer(f)
            w.writerow([bindings[param][len(BLDG) :] for param in params])


def pytest_generate_tests(metafunc):
    if metafunc.fixturenames == [
        "clean_building_motif",
        "template_name",
        "include_optional",
        "inline_dependencies",
    ]:
        test_cases = {
            "NOdep-NOoptional-NOinline": ("supply-fan", False, False),
            "NOdep-WITHoptional-NOinline": ("outside-air-damper", True, False),
            "WITHdep-WITHoptional-NOinline": ("single-zone-vav-ahu", True, False),
            "WITHdep-WITHoptional-WITHinline": ("single-zone-vav-ahu", True, True),
        }
        metafunc.parametrize(
            "template_name,include_optional,inline_dependencies",
            test_cases.values(),
            ids=test_cases.keys(),
        )


def test_template_generation_inmemory(
    clean_building_motif, template_name, include_optional, inline_dependencies
):
    fixture_lib = Library.load(directory="tests/unit/fixtures/templates")
    template = fixture_lib.get_template_by_name(template_name)
    template = fixture_lib.get_template_by_name(template_name)
    if inline_dependencies:
        template = template.inline_dependencies()
    bindings, filled = template.fill(BLDG, include_optional=include_optional)

    with NamedTemporaryFile(suffix=".xlsx", delete=False) as dest:
        output = template.generate_spreadsheet()
        assert output is not None
        dest.write(output.getbuffer())
        dest.flush()
        w = openpyxl.load_workbook(dest.name)
        _add_spreadsheet_row(w.active, bindings)
        w.save(Path(dest.name))

        reader = XLSXIngress(Path(dest.name))
        ing = TemplateIngress(template, None, reader)
        g = ing.graph(BLDG)

        assert isomorphic(
            g, filled
        ), f"Template -> spreadsheet -> ingress -> graph path did not generate a result isomorphic to just filling the template {template.name}"


def test_template_generation_file(
    clean_building_motif, template_name, include_optional, inline_dependencies
):
    fixture_lib = Library.load(directory="tests/unit/fixtures/templates")
    template = fixture_lib.get_template_by_name(template_name)
    if inline_dependencies:
        template = template.inline_dependencies()
    bindings, filled = template.fill(BLDG, include_optional=include_optional)

    with NamedTemporaryFile(suffix=".xlsx", delete=False) as dest:
        output = template.generate_spreadsheet(Path(dest.name))
        assert output is None
        dest.flush()
        w = openpyxl.load_workbook(dest.name)
        _add_spreadsheet_row(w.active, bindings)
        w.save(Path(dest.name))

        reader = XLSXIngress(Path(dest.name))
        ing = TemplateIngress(template, None, reader)
        g = ing.graph(BLDG)

        assert isomorphic(
            g, filled
        ), f"Template -> spreadsheet -> ingress -> graph path did not generate a result isomorphic to just filling the template {template.name}"


def test_csv_generation_inmemory(
    clean_building_motif, template_name, include_optional, inline_dependencies
):
    fixture_lib = Library.load(directory="tests/unit/fixtures/templates")
    template = fixture_lib.get_template_by_name(template_name)
    if inline_dependencies:
        template = template.inline_dependencies()
    bindings, filled = template.fill(BLDG, include_optional=include_optional)

    with NamedTemporaryFile(mode="w", suffix=".csv", delete=False) as dest:
        output = template.generate_csv()
        assert output is not None
        dest.writelines([output.getvalue()])
        dest.flush()

        params = output.getvalue().split(",")
        _add_csv_row(params, dest, bindings)

        reader = CSVIngress(Path(dest.name))
        ing = TemplateIngress(template, None, reader)
        g = ing.graph(BLDG)

        assert isomorphic(
            g, filled
        ), f"Template -> csv -> ingress -> graph path did not generate a result isomorphic to just filling the template {template.name}\n{(filled - g).serialize()}"


def test_csv_generation_file(
    clean_building_motif, template_name, include_optional, inline_dependencies
):
    fixture_lib = Library.load(directory="tests/unit/fixtures/templates")
    template = fixture_lib.get_template_by_name(template_name)
    if inline_dependencies:
        template = template.inline_dependencies()
    bindings, filled = template.fill(BLDG, include_optional=include_optional)

    with NamedTemporaryFile(mode="w", suffix=".csv", delete=False) as dest:
        output = template.generate_csv(Path(dest.name))
        assert output is None
        dest.flush()

        with open(Path(dest.name)) as f:
            params = f.read().strip().split(",")
        _add_csv_row(params, dest, bindings)

        reader = CSVIngress(Path(dest.name))
        ing = TemplateIngress(template, None, reader)
        g = ing.graph(BLDG)

        assert isomorphic(
            g, filled
        ), f"Template -> csv -> ingress -> graph path did not generate a result isomorphic to just filling the template {template.name}\n{(filled - g).serialize()}"
