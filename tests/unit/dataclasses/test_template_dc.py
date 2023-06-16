import csv
from pathlib import Path
from tempfile import NamedTemporaryFile

import openpyxl
import pytest
import rdflib
from rdflib import RDF, Namespace, URIRef
from rdflib.compare import isomorphic
from rdflib.namespace import FOAF
from sqlalchemy.exc import IntegrityError, NoResultFound

from buildingmotif.dataclasses import Library, Template
from buildingmotif.dataclasses.template import Dependency
from buildingmotif.ingresses.csv import CSVIngress
from buildingmotif.ingresses.template import TemplateIngress
from buildingmotif.ingresses.xlsx import XLSXIngress
from buildingmotif.template_compilation import compile_template_spec
from buildingmotif.utils import graph_size

BLDG = Namespace("urn:bldg/")

dependant_template_body = rdflib.Graph()
dependant_template_body.parse(
    data="""
@prefix P: <urn:___param___#> .
@prefix brick: <https://brickschema.org/schema/Brick#> .
P:name a brick:VAV ;
    brick:hasPoint P:1, P:2, P:3, P:4 .
"""
)

dependency_template_body = rdflib.Graph()
dependency_template_body.parse(
    data="""
@prefix P: <urn:___param___#> .
@prefix brick: <https://brickschema.org/schema/Brick#> .
P:name a brick:Temperature_Sensor ;
    brick:hasUnit P:param .
"""
)


def test_create(clean_building_motif):
    lib = Library.create("my_library")
    t = lib.create_template("my_template")

    assert isinstance(t, Template)
    assert isinstance(t.id, int)
    assert t.name == "my_template"
    assert isinstance(t.body, rdflib.Graph)
    assert t.defining_library == lib

    also_t = lib.get_templates()[0]
    assert also_t.id == t.id
    assert also_t.name == t.name
    assert isomorphic(also_t.body, t.body)
    assert also_t.defining_library == lib


def test_load(clean_building_motif):
    lib = Library.create("my_library")
    t = lib.create_template("my_template")
    t.body.add((URIRef("http://example.org/alex"), RDF.type, FOAF.Person))

    result = Template.load(t.id)
    assert result.id == t.id
    assert result.name == t.name
    assert isomorphic(result.body, t.body)


def test_set_name(clean_building_motif):
    lib = Library.create("my_library")
    t = lib.create_template("my_template")

    t.name = "new name"
    assert t.name == "new name"


def test_update_id(clean_building_motif):
    lib = Library.create("my_library")
    t = lib.create_template("my_template")

    with pytest.raises(AttributeError):
        t.id = 1


def test_save_body(clean_building_motif):
    lib = Library.create("my_library")
    t = lib.create_template("my_template")

    assert isinstance(t, Template)
    assert isinstance(t.id, int)
    assert t.name == "my_template"
    assert isinstance(t.body, rdflib.Graph)

    triple = (URIRef("http://example.org/alex"), RDF.type, FOAF.Person)
    t.body.add(triple)

    also_t = Template.load(t.id)
    assert also_t.id == t.id
    assert also_t.name == t.name
    assert isomorphic(also_t.body, t.body)


def test_add_dependency(clean_building_motif):
    lib = Library.create("my_library")
    dependant = lib.create_template("dependant", dependant_template_body)
    dependee = lib.create_template("dependee", dependency_template_body)

    dependant.add_dependency(dependee, {"name": "1", "param": "2"})

    assert dependant.get_dependencies() == (
        Dependency(dependee.id, {"name": "1", "param": "2"}),
    )
    dependant.check_dependencies()


def test_add_multiple_dependencies(clean_building_motif):
    lib = Library.create("my_library")
    dependant = lib.create_template("dependant", dependant_template_body)
    dependee = lib.create_template("dependee", dependency_template_body)

    dependant.add_dependency(dependee, {"name": "1", "param": "2"})
    dependant.add_dependency(dependee, {"name": "3", "param": "4"})

    assert (
        Dependency(dependee.id, {"name": "1", "param": "2"})
        in dependant.get_dependencies()
    )
    assert (
        Dependency(dependee.id, {"name": "3", "param": "4"})
        in dependant.get_dependencies()
    )
    assert len(dependant.get_dependencies()) == 2
    dependant.check_dependencies()


def test_add_dependency_bad_args(clean_building_motif):
    lib = Library.create("my_library")
    dependant = lib.create_template("dependant", dependant_template_body)
    dependee = lib.create_template("dependee", dependency_template_body)

    with pytest.raises(ValueError):
        dependant.add_dependency(dependee, {"bad": "xyz"})
        dependant.check_dependencies()


def test_add_dependency_already_exist(clean_building_motif):
    lib = Library.create("my_library")
    dependant = lib.create_template("dependant", dependant_template_body)
    dependee = lib.create_template("dependee", dependency_template_body)

    dependant.add_dependency(dependee, {"name": "1", "param": "2"})

    with pytest.raises(IntegrityError):
        dependant.add_dependency(dependee, {"name": "1", "param": "2"})

    clean_building_motif.session.rollback()


def test_get_dependencies(clean_building_motif):
    lib = Library.create("my_library")
    dependant = lib.create_template("dependant", dependant_template_body)
    dependee = lib.create_template("dependee", dependency_template_body)

    dependant.add_dependency(dependee, {"name": "1", "param": "2"})

    assert dependant.get_dependencies() == (
        Dependency(dependee.id, {"name": "1", "param": "2"}),
    )


def test_remove_dependency(clean_building_motif):
    lib = Library.create("my_library")
    dependant = lib.create_template("dependant", dependant_template_body)
    dependee = lib.create_template("dependee", dependency_template_body)

    dependant.add_dependency(dependee, {"name": "1", "param": "2"})
    assert dependant.get_dependencies() == (
        Dependency(dependee.id, {"name": "1", "param": "2"}),
    )

    dependant.remove_dependency(dependee)
    assert dependant.get_dependencies() == ()


def test_remove_depedancy_does_not_exist(clean_building_motif):
    lib = Library.create("my_library")
    dependant = lib.create_template("dependant", dependant_template_body)
    dependee = lib.create_template("dependee", dependency_template_body)

    with pytest.raises(NoResultFound):
        dependant.remove_dependency(dependee)

    clean_building_motif.session.rollback()


def test_get_library_dependencies(clean_building_motif):
    Library.load(ontology_graph="tests/unit/fixtures/Brick1.3rc1-equip-only.ttl")
    lib = Library.load(directory="tests/unit/fixtures/sample-lib-1")
    sf_templ = lib.get_template_by_name("fan")
    libs = sf_templ.library_dependencies()
    assert len(libs) == 2
    assert all(map(lambda x: isinstance(x, Library), libs))
    assert {str(lib.name) for lib in libs} == {
        "sample-lib-1",
        "https://brickschema.org/schema/1.3/Brick",
    }


def test_template_compilation(clean_building_motif):
    spec = {
        "hasPoint": {
            "occ": "https://brickschema.org/schema/Brick#Occupancy_Sensor",
            "temp": "https://brickschema.org/schema/Brick#Temperature_Sensor",
            "sp": "https://brickschema.org/schema/Brick#Temperature_Setpoint",
        },
        "downstream": {"zone": "https://brickschema.org/schema/Brick#HVAC_Zone"},
    }
    spec = compile_template_spec(spec)
    assert isinstance(spec, dict)
    assert isinstance(spec["body"], rdflib.Graph)
    spec["name"] = "test"
    lib = Library.create("my_library")
    # no dependencies to resolve, so we can just throw this away
    _ = spec.pop("dependencies")
    spec["optional_args"] = spec.pop("optional", [])
    templ = lib.create_template(**spec)
    assert templ.name == "test"
    assert sorted(templ.parameters) == sorted(("name", "occ", "temp", "sp", "zone"))
    assert graph_size(templ.body) == 8


# utility function for spreadsheet tests
def _add_spreadsheet_row(sheet, bindings):
    num_columns = sheet.max_column
    for column in range(1, num_columns + 1):
        param = sheet.cell(row=1, column=column).value
        sheet.cell(row=2, column=column).value = bindings[param][len(BLDG) :]


def _add_csv_row(output_io, tempfile, bindings):
    params = [param.strip() for param in output_io.getvalue().split(",")]
    w = csv.writer(tempfile)
    w.writerow([bindings[param][len(BLDG) :] for param in params])
    tempfile.flush()


def test_template_spreadsheet_generation_simple_nodeps_noopt_noinline(
    clean_building_motif,
):
    lib = Library.load(directory="tests/unit/fixtures/templates")

    # test simple template, no deps, no optional, no inline
    templ = lib.get_template_by_name("supply-fan")
    bindings, filled = templ.fill(BLDG)
    with NamedTemporaryFile(suffix=".xlsx") as dest:
        output = templ.generate_spreadsheet()
        assert output is not None
        dest.write(output.getbuffer())

        w = openpyxl.load_workbook(dest.name)
        _add_spreadsheet_row(w.active, bindings)
        w.save(Path(dest.name))

        reader = XLSXIngress(Path(dest.name))
        ing = TemplateIngress(templ, None, reader)
        g = ing.graph(BLDG)

        assert isomorphic(
            g, filled
        ), f"Template -> spreadsheet -> ingress -> graph path did not generate a result isomorphic to just filling the template {templ.name}"


def test_template_csv_generation_simple_nodeps_noopt_noinline(clean_building_motif):
    lib = Library.load(directory="tests/unit/fixtures/templates")

    # test simple template, no deps, no optional, no inline
    templ = lib.get_template_by_name("supply-fan")
    bindings, filled = templ.fill(BLDG)
    with NamedTemporaryFile(mode="w", suffix=".csv") as dest:
        output = templ.generate_csv()
        assert output is not None
        dest.writelines([output.getvalue()])
        dest.flush()

        _add_csv_row(output, dest, bindings)

        reader = CSVIngress(Path(dest.name))
        ing = TemplateIngress(templ, None, reader)
        g = ing.graph(BLDG)

        assert isomorphic(
            g, filled
        ), f"Template -> csv -> ingress -> graph path did not generate a result isomorphic to just filling the template {templ.name}"


def test_template_spreadsheet_generation_simple_nodeps_withopt_noinline(
    clean_building_motif,
):
    lib = Library.load(directory="tests/unit/fixtures/templates")

    # test simple template, no deps, WITH optional, no inline
    templ = lib.get_template_by_name("outside-air-damper")
    bindings, filled = templ.fill(BLDG, include_optional=True)
    with NamedTemporaryFile(suffix=".xlsx") as dest:
        output = templ.generate_spreadsheet()
        assert output is not None
        dest.write(output.getbuffer())

        w = openpyxl.load_workbook(dest.name)
        _add_spreadsheet_row(w.active, bindings)
        w.save(Path(dest.name))

        reader = XLSXIngress(Path(dest.name))
        ing = TemplateIngress(templ, None, reader)
        g = ing.graph(BLDG)

        assert isomorphic(
            g, filled
        ), f"Template -> spreadsheet -> ingress -> graph path did not generate a result isomorphic to just filling the template {templ.name}"


def test_template_csv_generation_simple_nodeps_withopt_noinline(clean_building_motif):
    lib = Library.load(directory="tests/unit/fixtures/templates")

    # test simple template, no deps, WITH optional, no inline
    templ = lib.get_template_by_name("outside-air-damper")
    bindings, filled = templ.fill(BLDG, include_optional=True)
    with NamedTemporaryFile(mode="w", suffix=".csv") as dest:
        output = templ.generate_csv()
        assert output is not None
        dest.writelines([output.getvalue()])
        dest.flush()

        _add_csv_row(output, dest, bindings)

        reader = CSVIngress(Path(dest.name))
        ing = TemplateIngress(templ, None, reader)
        g = ing.graph(BLDG)

        assert isomorphic(
            g, filled
        ), f"Template -> csv -> ingress -> graph path did not generate a result isomorphic to just filling the template {templ.name}"


def test_template_spreadsheet_generation_simple_withdeps_withopt_noinline(
    clean_building_motif,
):
    lib = Library.load(directory="tests/unit/fixtures/templates")

    # test simple template, WITH deps, WITH optional, no inline
    templ = lib.get_template_by_name("single-zone-vav-ahu")
    bindings, filled = templ.fill(BLDG, include_optional=True)
    with NamedTemporaryFile(suffix=".xlsx") as dest:
        output = templ.generate_spreadsheet()
        assert output is not None
        dest.write(output.getbuffer())

        w = openpyxl.load_workbook(dest.name)
        _add_spreadsheet_row(w.active, bindings)
        w.save(Path(dest.name))

        reader = XLSXIngress(Path(dest.name))
        ing = TemplateIngress(templ, None, reader)
        g = ing.graph(BLDG)

        assert isomorphic(
            g, filled
        ), f"Template -> spreadsheet -> ingress -> graph path did not generate a result isomorphic to just filling the template {templ.name}"


def test_template_csv_generation_simple_withdeps_withopt_noinline(clean_building_motif):
    lib = Library.load(directory="tests/unit/fixtures/templates")

    # test simple template, WITH deps, WITH optional, no inline
    templ = lib.get_template_by_name("single-zone-vav-ahu")
    bindings, filled = templ.fill(BLDG, include_optional=True)
    with NamedTemporaryFile(mode="w", suffix=".csv") as dest:
        output = templ.generate_csv()
        assert output is not None
        dest.writelines([output.getvalue()])
        dest.flush()

        _add_csv_row(output, dest, bindings)

        reader = CSVIngress(Path(dest.name))
        ing = TemplateIngress(templ, None, reader)
        g = ing.graph(BLDG)

        assert isomorphic(
            g, filled
        ), f"Template -> csv -> ingress -> graph path did not generate a result isomorphic to just filling the template {templ.name}"


def test_template_spreadsheet_generation_simple_withdeps_withopt_withinline(
    clean_building_motif,
):
    lib = Library.load(directory="tests/unit/fixtures/templates")

    # test simple template, WITH deps, WITH optional, WITH inline
    templ = lib.get_template_by_name("single-zone-vav-ahu").inline_dependencies()
    bindings, filled = templ.fill(BLDG, include_optional=True)
    with NamedTemporaryFile(suffix=".xlsx") as dest:
        output = templ.generate_spreadsheet()
        assert output is not None
        dest.write(output.getbuffer())

        w = openpyxl.load_workbook(dest.name)
        _add_spreadsheet_row(w.active, bindings)
        w.save(Path(dest.name))

        reader = XLSXIngress(Path(dest.name))
        ing = TemplateIngress(templ, None, reader, inline=True)
        g = ing.graph(BLDG)

        assert isomorphic(
            g, filled
        ), f"Template -> spreadsheet -> ingress -> graph path did not generate a result isomorphic to just filling the template {templ.name}"


def test_template_csv_generation_simple_withdeps_withopt_withinline(
    clean_building_motif,
):
    lib = Library.load(directory="tests/unit/fixtures/templates")

    # test simple template, WITH deps, WITH optional, WITH inline
    templ = lib.get_template_by_name("single-zone-vav-ahu").inline_dependencies()
    bindings, filled = templ.fill(BLDG, include_optional=True)
    with NamedTemporaryFile(mode="w", suffix=".csv") as dest:
        output = templ.generate_csv()
        assert output is not None
        dest.writelines([output.getvalue()])
        dest.flush()

        _add_csv_row(output, dest, bindings)

        reader = CSVIngress(Path(dest.name))
        ing = TemplateIngress(templ, None, reader, inline=True)
        g = ing.graph(BLDG)

        assert isomorphic(
            g, filled
        ), f"Template -> csv -> ingress -> graph path did not generate a result isomorphic to just filling the template {templ.name}"
